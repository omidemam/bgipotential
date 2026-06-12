# SENSITIVITY ANALYSIS — 2-Simplex Grid Search with Spearman rho vs. Base Weights
# This script is a sensitivity analysis of a composite index called the Green Infrastructure Potential Index,
# quantifying the potential of green infrastructure to mitigate urban heat island effect, flooding, and air pollution.
#
# Ternary plots use filled contours (tricontourf) for a smooth topographic appearance,
# matching the style of the reference image (colored iso-bands + white isoline overlay).

import numpy as np
import pandas as pd
import os
import matplotlib.pyplot as plt
import matplotlib.tri as mtri
import matplotlib.patheffects as pe
from matplotlib.patches import PathPatch
from matplotlib.path import Path

# ─────────────────────────────────────────────
# USER SETTINGS — edit these before running
# ─────────────────────────────────────────────
SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT    = os.path.abspath(os.path.join(SCRIPT_DIR, "..", ".."))
FILEPATH     = os.path.join(REPO_ROOT, "Results", "GIPI_Results.xls")
BASE_WEIGHTS = np.array([0.60, 0.30, 0.10])   # HydroWeight, HeatWeight, AQWeight
N_SIMS       = 5000                             # more points = smoother surface
RANDOM_SEED  = 35
OUTPUT_DIR   = os.path.join(REPO_ROOT, "Visualization", "Figures and maps")
OUTPUT_COMBINED = os.path.join(OUTPUT_DIR, "GIPI_simplex_combined2.png")

# Cities to show in the ridgeline joy plot (panel c)
RIDGELINE_CITIES = ["NYC", "Charlotte", "Denver", "SanJose"]

# ── Optional export paths (used in STEP 11 & STEP 12 below) ──────────────
OUTPUT_STABILITY_XLSX = os.path.join(OUTPUT_DIR, "GIPI_RankStability_5_5.xlsx")
OUTPUT_RANKDIST_IMAGE = os.path.join(OUTPUT_DIR, "GIPI_RankDistributions_5_5.png")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Column definitions
HYDRO_COLS = ["HydroRiskAvg"]
HEAT_COLS  = ["HeatRiskAvg"]
AQ_COLS    = ["AQAvg"]


# ─────────────────────────────────────────────
# STEP 1 — Load and prepare data
# ─────────────────────────────────────────────
def load_and_prepare_data(filepath, hydro_cols, heat_cols, aq_cols):
    """
    Load a CSV or Excel file (.csv, .xls, .xlsx) and aggregate raw columns
    into 3 fixed category scores per city.
    pandas supports both formats equally well:
      - CSV  → pd.read_csv()
      - Excel → pd.read_excel()  (requires the 'xlrd' package for .xls)
    """
    ext = filepath.rsplit(".", 1)[-1].lower()
    if ext in ("xls", "xlsx"):
        data = pd.read_excel(filepath)
    else:
        data = pd.read_csv(filepath)
    city_names = data["City"].values

    hydro_score = data[hydro_cols].mean(axis=1).values
    heat_score  = data[heat_cols].mean(axis=1).values
    aq_score    = data[aq_cols].mean(axis=1).values

    X = np.column_stack([hydro_score, heat_score, aq_score])  # shape: (n_cities, 3)
    return city_names, X


# ─────────────────────────────────────────────
# STEP 2 — Compute rank vector for a given weight set
# ─────────────────────────────────────────────
def rank_cities(X, weights):
    """Score and rank cities (1 = highest score)."""
    scores = X @ weights
    return np.argsort(np.argsort(-scores)) + 1


# ─────────────────────────────────────────────
# STEP 3 — Sample weight schemes uniformly over the 2-simplex
# ─────────────────────────────────────────────
def sample_simplex(n_sims, rng):
    """
    Draw n_sims weight triplets (w1, w2, w3) UNIFORMLY from the 2-simplex
    using Dirichlet(1, 1, 1) sampling — the mathematically correct method
    for uniform coverage of the simplex surface.

    Note: The broken-stick approach (w1~U(0,1), w2~U(0,1-w1), w3=remainder)
    is biased — it over-samples near the w1≈0 vertex because w1 gets
    unrestricted range while w2 and w3 must share what remains.
    Dirichlet(1,1,1) avoids this bias by treating all three weights symmetrically.
    """
    return rng.dirichlet([1, 1, 1], size=n_sims)   # shape: (n_sims, 3)


# ─────────────────────────────────────────────
# STEP 4 — Run grid search and compute Spearman rho vs. base ranking
# ─────────────────────────────────────────────
def run_grid_search(X, base_weights, weight_schemes, city_names):
    """
    For each sampled weight scheme:
      1. Rank all cities.
      2. Compute Spearman rank correlation between that ranking and the
         base-weight ranking across all cities.
      3. Record every city's rank for that trial (used for rank distributions).
    Returns:
      results      — DataFrame of weight schemes + Spearman rho, sorted descending.
      all_ranks_df — DataFrame (n_sims × n_cities) of per-trial city ranks.
    """
    base_ranks = rank_cities(X, base_weights)
    n_cities   = X.shape[0]

    records   = []
    rank_rows = []          # one list of ranks per trial

    for w in weight_schemes:
        w1, w2, w3 = w
        trial_ranks = rank_cities(X, w)

        rho = np.corrcoef(base_ranks, trial_ranks)[0, 1]

        # Calculate the mean absolute shift in rank across all cities
        mean_rank_shift = np.mean(np.abs(trial_ranks - base_ranks))

        records.append({
            "w_hydro"       : round(w1, 4),
            "w_heat"        : round(w2, 4),
            "w_aq"          : round(w3, 4),
            "SpearmanRho"   : round(rho, 6),
            "MeanRankShift" : round(mean_rank_shift, 4),
        })
        rank_rows.append(trial_ranks)

    results      = pd.DataFrame(records)
    all_ranks_df = pd.DataFrame(rank_rows, columns=city_names)

    return results.sort_values("SpearmanRho", ascending=False).reset_index(drop=True), all_ranks_df


# ─────────────────────────────────────────────
# STEP 5 — Ternary plot helpers (shared)
# ─────────────────────────────────────────────
def simplex_to_cartesian(w1, w2, w3):
    """
    Project 3D simplex coordinates to 2D Cartesian for a ternary plot.

    Vertex layout (matches axis labels below):
      Bottom-left  = Hydro  (w1=1, w2=0, w3=0) → (0, 0)
      Bottom-right = Heat   (w1=0, w2=1, w3=0) → (1, 0)
      Top          = AQ     (w1=0, w2=0, w3=1) → (0.5, √3/2)
    """
    x = w2 + w3 * 0.5
    y = w3 * (np.sqrt(3) / 2)
    return x, y


def _triangle_clip_patch():
    """
    Return a PathPatch whose boundary is the simplex triangle.
    Used to clip tricontourf fills to exactly the triangle interior.
    Vertices: (0,0) → (1,0) → (0.5, √3/2) → closed.
    """
    h = np.sqrt(3) / 2
    verts = [(0, 0), (1, 0), (0.5, h), (0, 0)]
    codes = [Path.MOVETO, Path.LINETO, Path.LINETO, Path.CLOSEPOLY]
    return PathPatch(Path(verts, codes), transform=None)


def _add_grid_and_labels(ax):
    """
    Draw iso-weight dashed grid lines and percentage tick labels.
    Lines are drawn in white so they remain visible over the contour fill.
    """
    tick_levels = [0.2, 0.4, 0.6, 0.8]
    text_props  = dict(color="#555", fontsize=14, ha="center", va="center")

    for level in tick_levels:
        # 1. Lines of constant w1 (Hydro) — bottom edge → left edge
        p1 = np.array(simplex_to_cartesian(level, 1 - level, 0))
        p2 = np.array(simplex_to_cartesian(level, 0, 1 - level))
        ax.plot([p1[0], p2[0]], [p1[1], p2[1]],
                color="white", alpha=0.55, lw=1.0, zorder=4, ls="--")
        off1 = np.array([-np.sqrt(3)/2, 0.5]) * 0.07
        ax.text(p2[0] + off1[0], p2[1] + off1[1],
                f"{int(level*100)}%", rotation=60, **text_props)

        # 2. Lines of constant w2 (Heat) — bottom edge → right edge
        p1 = np.array(simplex_to_cartesian(1 - level, level, 0))
        p2 = np.array(simplex_to_cartesian(0, level, 1 - level))
        ax.plot([p1[0], p2[0]], [p1[1], p2[1]],
                color="white", alpha=0.55, lw=1.0, zorder=4, ls="--")
        off2 = np.array([0, -1]) * 0.07
        ax.text(p1[0] + off2[0], p1[1] + off2[1],
                f"{int(level*100)}%", rotation=0, **text_props)

        # 3. Lines of constant w3 (AQ) — right edge → left edge
        p1 = np.array(simplex_to_cartesian(0, 1 - level, level))
        p2 = np.array(simplex_to_cartesian(1 - level, 0, level))
        ax.plot([p1[0], p2[0]], [p1[1], p2[1]],
                color="white", alpha=0.55, lw=1.0, zorder=4, ls="--")
        off3 = np.array([np.sqrt(3)/2, 0.5]) * 0.07
        ax.text(p1[0] + off3[0], p1[1] + off3[1],
                f"{int(level*100)}%", rotation=-60, **text_props)


def _add_vertex_labels(ax):
    """Place bold axis-corner labels just outside each vertex of the triangle."""
    offset = 0.09
    label_style = dict(
        ha="center", va="center", fontsize=18,
        color="#1a1a1a", fontweight="bold",
        path_effects=[pe.withStroke(linewidth=4, foreground="white")]
    )
    ax.text(0 - offset, 0 - offset * 0.8, "Hydro\n(w₁=1)", **label_style)
    ax.text(1 + offset, 0 - offset * 0.8, "Heat\n(w₂=1)", **label_style)
    ax.text(0.5,  np.sqrt(3)/2 + offset,  "AQ\n(w₃=1)",   **label_style)


def _base_weight_annotation(ax, base_weights, color, shift_label=False):
    """Plot the base-weight star marker and annotation arrow."""
    bx, by = simplex_to_cartesian(*base_weights)
    label = (
        f"Base\n({base_weights[0]}, {base_weights[1]}, {base_weights[2]})\nShift=0"
        if shift_label else
        f"Base\n({base_weights[0]}, {base_weights[1]}, {base_weights[2]})"
    )
    ax.scatter(bx, by, marker="*", s=450, color=color, zorder=8,
               edgecolors="#333", linewidths=1.0,
               label=f"Base weights ({base_weights[0]}, {base_weights[1]}, {base_weights[2]})")
    ax.annotate(
        label,
        xy=(bx, by), xytext=(bx + 0.07, by + 0.04),
        color=color, fontsize=14, fontweight="bold",
        arrowprops=dict(arrowstyle="->", color=color, lw=1.2),
        path_effects=[pe.withStroke(linewidth=3, foreground="white")],
        zorder=9,
    )


# ─────────────────────────────────────────────
# STEP 5a+5b — Combined ternary contour plot
# ─────────────────────────────────────────────
def _draw_ternary_panel(ax, triang, values, levels, cmap, cbar_label,
                        base_weights, bw_color, shift_label, fig):
    """Draw one ternary contour panel on the given axes."""
    # ── Filled contours ───────────────────────────────────────────────────
    cf = ax.tricontourf(triang, values, levels=levels,
                        cmap=cmap, zorder=2, antialiased=True)
    ax.tricontour(triang, values, levels=levels,
                  colors="white", linewidths=0.5, alpha=0.45, zorder=3)

    # ── Clip to triangle ──────────────────────────────────────────────────
    clip = _triangle_clip_patch()
    clip.set_transform(ax.transData)
    for coll in ax.collections:
        coll.set_clip_path(clip)

    # ── Triangle border ───────────────────────────────────────────────────
    h = np.sqrt(3) / 2
    tri_verts = np.array([[0, 0], [1, 0], [0.5, h], [0, 0]])
    ax.plot(tri_verts[:, 0], tri_verts[:, 1], color="#2b2b2b", lw=1.8, zorder=5)

    # ── Grid, vertex labels, base-weight marker ───────────────────────────
    _add_grid_and_labels(ax)
    _add_vertex_labels(ax)
    _base_weight_annotation(ax, base_weights, color=bw_color, shift_label=shift_label)

    # ── Colorbar ──────────────────────────────────────────────────────────
    cbar = fig.colorbar(cf, ax=ax, shrink=0.55, pad=0.02)
    cbar.set_label(cbar_label, color="#1a1a1a", fontsize=15, labelpad=10)
    cbar.ax.yaxis.set_tick_params(color="#1a1a1a")
    plt.setp(cbar.ax.yaxis.get_ticklabels(), color="#1a1a1a", fontsize=12)
    cbar.outline.set_edgecolor("#999")

    # ── Legend ─────────────────────────────────────────────────────────────
    ax.legend(loc="lower center", fontsize=12, framealpha=0.8,
              labelcolor="#1a1a1a", facecolor="white", edgecolor="#ccc")


# ─────────────────────────────────────────────
# STEP 5c — Ridgeline (joy plot) panel helper
# ─────────────────────────────────────────────
def plot_ridgeline_panel(ax, all_ranks_df, city_names_all, ridgeline_cities,
                         X, base_weights):
    """
    Draw a ridgeline / joy-plot of rank distributions for the specified cities
    on a single shared set of axes.  Each city gets its own color; a dashed
    vertical line (same color) marks that city's Classic GIPI rank.
    The KDE is forced to zero outside [1, 20].
    """
    from scipy.stats import gaussian_kde

    # ── Resolve city names (case-insensitive) ─────────────────────────────
    matched = []
    for city in ridgeline_cities:
        m = city if city in all_ranks_df.columns else next(
            (c for c in all_ranks_df.columns if c.lower() == city.lower()), None
        )
        if m is None:
            print(f"[Ridgeline WARNING] '{city}' not found — skipping.")
        else:
            matched.append(m)

    if not matched:
        return

    # ── Display names: map raw column names → pretty legend labels ───────────
    DISPLAY_NAMES = {
        "SanJose"       : "San Jose",
        "Charlotte"     : "Charlotte",
        "Denver"        : "Denver",
        "NYC"           : "NYC",
    }
    def _display(name):
        return DISPLAY_NAMES.get(name, name)

    # ── Classic GIPI ranks ────────────────────────────────────────────────
    base_rank_vec = rank_cities(X, base_weights)
    classic_ranks = {name: int(base_rank_vec[i]) for i, name in enumerate(city_names_all)}

    # ── Color palette — one per city ──────────────────────────────────────
    PALETTE = ["#5b9bd5", "#e05c5c", "#8e6bbf", "#4cae8a",
               "#e8a838", "#d45f8a", "#3a9e9e", "#b07d4a"]
    city_colors = {city: PALETTE[i % len(PALETTE)] for i, city in enumerate(matched)}

    # ── Evaluation grid (fine, same for every city) ───────────────────────
    x_eval = np.linspace(0.5, 20.5, 1000)
    mask   = (x_eval >= 1) & (x_eval <= 20)   # taper to 0 outside [1,20]

    ax.set_facecolor("white")

    legend_handles = []

    for city in matched:
        color     = city_colors[city]
        samples   = all_ranks_df[city].values.astype(float)
        kde       = gaussian_kde(samples, bw_method="scott")
        density   = kde(x_eval)
        density   = np.where(mask, density, 0.0)   # zero outside [1,20]

        # Normalise so area under [1,20] sums to 1
        # (KDE is already a proper density; clipping can change the integral
        #  slightly, so we just leave it — the shape is what matters)

        line, = ax.plot(x_eval, density, color=color, linewidth=2.0,
                        label=_display(city), zorder=3)
        ax.fill_between(x_eval, density, alpha=0.18, color=color, zorder=2)
        legend_handles.append(line)

        # ── Classic GIPI rank dashed line ─────────────────────────────────
        classic_r = classic_ranks.get(city, None)
        if classic_r is not None:
            ax.axvline(x=classic_r, color=color, linewidth=1.6,
                       linestyle="--", zorder=4)
            # Label the rank value above the line
            y_top = kde(np.array([classic_r]))[0]
            ax.text(classic_r, y_top * 1.05, str(classic_r),
                    color=color, fontsize=10, ha="center", va="bottom",
                    fontweight="bold")

    # ── Axes formatting ───────────────────────────────────────────────────
    ax.set_xlim(0.5, 20.5)
    ax.set_ylim(bottom=0)
    ax.set_xticks(range(1, 21))
    ax.set_xlabel("Rank", fontsize=18, color="#1a1a1a", labelpad=6)
    ax.set_ylabel("Probability", fontsize=18, color="#1a1a1a", labelpad=6)
    ax.tick_params(colors="#1a1a1a", labelsize=14)
    for spine in ax.spines.values():
        spine.set_edgecolor("#cccccc")
    ax.yaxis.grid(True, color="#eeeeee", linewidth=0.7, zorder=0)
    ax.set_axisbelow(True)

    # ── Legend (includes city lines; dashed = GIPI rank is implicit) ──────
    # Add a phantom dashed entry to explain the dotted lines
    import matplotlib.lines as mlines
    dash_handle = mlines.Line2D([], [], color="grey", linestyle="--",
                                linewidth=1.6, label="Classic GIPI rank")
    ax.legend(
        handles=legend_handles + [dash_handle],
        loc="upper right", fontsize=15,
        framealpha=0.85, facecolor="white", edgecolor="#cccccc",
        title="City", title_fontsize=16,
    )


def plot_combined_ternary(results, base_weights, output_path,
                          all_ranks_df=None, city_names_all=None, X_data=None,
                          ridgeline_cities=None, n_levels=14):
    """
    Side-by-side ternary contour plot (rows a & b) with an optional ridgeline
    joy-plot panel (c) below spanning the full figure width.
    Left (a): Mean Absolute Rank Shift. Right (b): Spearman rho vs. base ranking.
    Bottom (c): Rank KDE distributions for focus cities.
    A gray vertical divider is drawn between the two top panels.
    """
    import matplotlib.gridspec as gridspec

    w1    = results["w_hydro"].values
    w2    = results["w_heat"].values
    w3    = results["w_aq"].values
    rho = results["SpearmanRho"].values
    shift = results["MeanRankShift"].values

    x, y = simplex_to_cartesian(w1, w2, w3)

    # ── Triangulation (shared) ────────────────────────────────────────────
    triang = mtri.Triangulation(x, y)
    cx = x[triang.triangles].mean(axis=1)
    cy = y[triang.triangles].mean(axis=1)
    tri_mask = ~(
        (cy >= 0) &
        (cy <= np.sqrt(3) * cx) &
        (cy <= np.sqrt(3) * (1 - cx))
    )
    triang.set_mask(tri_mask)

    # ── Decide whether to draw ridgeline row ──────────────────────────────
    draw_ridge = (
        all_ranks_df is not None and
        city_names_all is not None and
        X_data is not None and
        ridgeline_cities is not None
    )

    # ── Figure layout ─────────────────────────────────────────────────────
    if draw_ridge:
        fig = plt.figure(figsize=(20, 14), facecolor="white")
        # Two rows: top = ternary panels, bottom = ridgeline
        gs_outer = gridspec.GridSpec(
            2, 1,
            height_ratios=[1.1, 0.7],
            hspace=0.35,
            figure=fig,
        )
        gs_top = gridspec.GridSpecFromSubplotSpec(
            1, 3,
            subplot_spec=gs_outer[0],
            width_ratios=[1, 0.01, 1],
            wspace=0.28,
        )
        ax_ridge = fig.add_subplot(gs_outer[1])
    else:
        fig = plt.figure(figsize=(20, 9), facecolor="white")
        gs_top = gridspec.GridSpec(1, 3, width_ratios=[1, 0.01, 1],
                                   wspace=0.28, figure=fig)
        ax_ridge = None

    ax_left  = fig.add_subplot(gs_top[0, 0])
    ax_div   = fig.add_subplot(gs_top[0, 1])
    ax_right = fig.add_subplot(gs_top[0, 2])

    for ax in (ax_left, ax_right):
        ax.set_facecolor("white")
        ax.set_aspect("equal")
        ax.axis("off")

    # ── Gray divider between ternary panels ───────────────────────────────
    ax_div.set_xlim(0, 1)
    ax_div.set_ylim(0, 1)
    ax_div.axvline(x=0.5, color="#999999", linewidth=1.5)
    ax_div.axis("off")

    # ── Left panel (a): MARS ──────────────────────────────────────────────
    levels_shift = np.linspace(shift.min(), shift.max(), n_levels + 1)
    _draw_ternary_panel(ax_left, triang, shift, levels_shift,
                        cmap="RdYlGn_r",
                        cbar_label="Mean Abs. Rank Shift (vs. base)",
                        base_weights=base_weights, bw_color="#1a237e",
                        shift_label=True, fig=fig)

    # ── Right panel (b): Spearman rho ─────────────────────────────────────
    triang2 = mtri.Triangulation(x, y)
    triang2.set_mask(tri_mask)
    levels_rho = np.linspace(rho.min(), rho.max(), n_levels + 1)
    _draw_ternary_panel(ax_right, triang2, rho, levels_rho,
                        cmap="YlOrRd",
                        cbar_label="Spearman rank correlation, ρ\n(vs. base-weight ranking)",
                        base_weights=base_weights, bw_color="#1b5e20",
                        shift_label=False, fig=fig)

    # ── Panel labels (a) and (b) ──────────────────────────────────────────
    label_style = dict(fontsize=22, fontweight="bold", color="#1a1a1a",
                       ha="left", va="top",
                       transform=fig.transFigure)
    fig.text(0.07, 0.97, "(a)", **label_style)
    fig.text(0.52, 0.97, "(b)", **label_style)

    # ── Ridgeline panel (c) ───────────────────────────────────────────────
    if draw_ridge:
        plot_ridgeline_panel(
            ax_ridge, all_ranks_df, city_names_all,
            ridgeline_cities, X_data, base_weights,
        )
        # (c) label — positioned relative to the ridgeline axes in figure coords
        ax_ridge_pos = ax_ridge.get_position()   # may not be final before draw
        fig.text(0.02, ax_ridge_pos.y1 + 0.01, "(c)",
                 fontsize=22, fontweight="bold", color="#1a1a1a",
                 ha="left", va="bottom", transform=fig.transFigure)

    plt.savefig(output_path, dpi=180, bbox_inches="tight",
                facecolor=fig.get_facecolor())
    print(f"\nCombined ternary plot saved to: {output_path}")
    plt.show(block=False)
    plt.pause(1)
    plt.close()


# ─────────────────────────────────────────────
# STEP 6 — Summary statistics: rank std-dev + distributions
# ─────────────────────────────────────────────
FOCUS_CITIES = None   # None → use ALL cities in the dataset (alphabetical); or set to a list of names

def summary_statistics(city_names, all_ranks_df, focus_cities, X=None, base_weights=None):
    """
    Print two blocks of summary statistics derived from the per-trial rank matrix.

    Block A — Rank std-dev for every city
    ──────────────────────────────────────
    Higher std-dev → city's rank is more sensitive to weight choices.

    Block B — Rank probability distribution for each focused city
    ─────────────────────────────────────────────────────────────────
    For each rank position (1 … n_cities) shows:
      • count of trials where city held that rank
      • percentage probability
      • a simple bar chart
    Also reports the mean rank, median rank, std-dev, and the Classic GIPI rank.
    """
    n_cities = len(city_names)
    n_sims   = len(all_ranks_df)

    # ── Resolve focus list: None → all cities sorted alphabetically ───────
    if focus_cities is None:
        resolved_focus = sorted(all_ranks_df.columns.tolist())
    else:
        resolved_focus = focus_cities

    # ── Compute Classic GIPI rank (base weights) if provided ──────────────
    classic_ranks = {}   # city_name → classic rank (int)
    if X is not None and base_weights is not None:
        base_rank_vec = rank_cities(X, base_weights)   # shape (n_cities,)
        for i, city in enumerate(city_names):
            classic_ranks[city] = int(base_rank_vec[i])

    # ── Block A: Rank std-dev for all cities ──────────────────────────────
    print("\n" + "═" * 56)
    print("  RANK STABILITY — Std-Dev Across All Weight Schemes")
    print("═" * 56)
    print(f"  (based on {n_sims:,} sampled weight triplets)\n")
    print(f"  {'City':<25} {'Mean Rank':>10} {'Median Rank':>12} {'Rank Std-Dev':>13} {'Classic GIPI':>13}")
    print("  " + "-" * 76)

    rank_stats = []
    for city in city_names:
        col = all_ranks_df[city]
        rank_stats.append({
            "city"    : city,
            "mean"    : col.mean(),
            "median"  : col.median(),
            "std"     : col.std(),
            "classic" : classic_ranks.get(city, ""),
        })
    rank_stats_df = pd.DataFrame(rank_stats).sort_values("mean")

    for _, row in rank_stats_df.iterrows():
        classic_str = f"{int(row['classic']):>13}" if row['classic'] != "" else f"{'N/A':>13}"
        print(f"  {row['city']:<25} {row['mean']:>10.2f} {row['median']:>12.1f} {row['std']:>13.4f}{classic_str}")

    # ── Block B: Full rank distributions for focused cities ───────────
    bar_width = 30          # characters for the ASCII bar

    for city in resolved_focus:
        # Try case-insensitive match if exact name not found
        matched = city if city in all_ranks_df.columns else next(
            (c for c in all_ranks_df.columns if c.lower() == city.lower()), None
        )
        if matched is None:
            print(f"\n  [WARNING] '{city}' not found in dataset — skipping distribution.")
            continue

        col       = all_ranks_df[matched]
        mean_rank = col.mean()
        med_rank  = col.median()
        std_rank  = col.std()
        classic_r = classic_ranks.get(matched, None)

        print("\n" + "═" * 56)
        print(f"  RANK PROBABILITY DISTRIBUTION — {matched}")
        print("═" * 56)
        classic_label = f"   Classic GIPI rank: {classic_r}" if classic_r is not None else ""
        print(f"  Mean rank: {mean_rank:.2f}   Median: {med_rank:.1f}   Std-Dev: {std_rank:.4f}{classic_label}")
        print(f"  {'Rank':<6} {'Count':>6} {'Prob %':>8}   Bar")
        print("  " + "-" * 56)

        counts = col.value_counts().sort_index()
        for rank_pos in range(1, n_cities + 1):
            count = counts.get(rank_pos, 0)
            prob  = count / n_sims * 100
            bar   = "█" * int(round(prob / 100 * bar_width))
            classic_marker = "  ◄ Classic GIPI" if (classic_r is not None and rank_pos == classic_r) else ""
            print(f"  {rank_pos:<6} {count:>6,} {prob:>7.2f}%   {bar}{classic_marker}")


# ─────────────────────────────────────────────
# STEP 7 — Run pipeline
# ─────────────────────────────────────────────
rng = np.random.default_rng(RANDOM_SEED)

city_names, X  = load_and_prepare_data(FILEPATH, HYDRO_COLS, HEAT_COLS, AQ_COLS)
weight_schemes = sample_simplex(N_SIMS, rng)
results, all_ranks_df = run_grid_search(X, BASE_WEIGHTS, weight_schemes, city_names)

# ─────────────────────────────────────────────
# STEP 8 — Spearman rank-correlation summary
# ─────────────────────────────────────────────
print(f"\nBase weights — Hydro: {BASE_WEIGHTS[0]}, Heat: {BASE_WEIGHTS[1]}, AQ: {BASE_WEIGHTS[2]}  |  {N_SIMS:,} Dirichlet samples\n")
print(f"--- Spearman Rank Correlation Summary ---")
print(f"Highest rho:  {results['SpearmanRho'].max():.6f}  "
      f"(w_hydro={results.iloc[0]['w_hydro']}, "
      f"w_heat={results.iloc[0]['w_heat']}, "
      f"w_aq={results.iloc[0]['w_aq']})")
print(f"Lowest  rho:  {results['SpearmanRho'].min():.6f}  "
      f"(w_hydro={results.iloc[-1]['w_hydro']}, "
      f"w_heat={results.iloc[-1]['w_heat']}, "
      f"w_aq={results.iloc[-1]['w_aq']})")
print(f"Mean    rho:  {results['SpearmanRho'].mean():.6f}")
print(f"Median  rho:  {results['SpearmanRho'].median():.6f}")

print(f"\n--- Mean Absolute Rank Shift Summary ---")
print(f"Largest Shift:  {results['MeanRankShift'].max():.4f}")
print(f"Smallest Shift: {results['MeanRankShift'].min():.4f}")
print(f"Mean Shift:     {results['MeanRankShift'].mean():.4f}")
print(f"Median Shift:   {results['MeanRankShift'].median():.4f}")

# ─────────────────────────────────────────────
# STEP 9 — Extended rank statistics
# ─────────────────────────────────────────────
summary_statistics(city_names, all_ranks_df, FOCUS_CITIES, X=X, base_weights=BASE_WEIGHTS)

# ─────────────────────────────────────────────
# STEP 10 — Generate combined ternary contour plot
# ─────────────────────────────────────────────
plot_combined_ternary(
    results, BASE_WEIGHTS, OUTPUT_COMBINED,
    all_ranks_df=all_ranks_df,
    city_names_all=city_names,
    X_data=X,
    ridgeline_cities=RIDGELINE_CITIES,
)


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  STEP 11 — OPTIONAL EXPORT: Rank Stability Table → Excel               ║
# ║  To disable: comment out everything from this banner to END STEP 11    ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def export_rank_stability_xlsx(city_names, all_ranks_df, output_path):
    """
    Build the rank-stability summary table (mean, median, std-dev per city,
    sorted by mean rank) and write it to an Excel spreadsheet.
    """
    rank_stats = []
    n_sims = len(all_ranks_df)
    for city in city_names:
        col = all_ranks_df[city]
        rank_stats.append({
            "City"        : city,
            "Mean Rank"   : round(col.mean(),   4),
            "Median Rank" : round(col.median(), 4),
            "Rank Std-Dev": round(col.std(),    4),
        })
    df_out = pd.DataFrame(rank_stats).sort_values("Mean Rank").reset_index(drop=True)
    df_out.index += 1   # 1-based row numbers

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_out.to_excel(writer, sheet_name="Rank Stability", index=True, index_label="#")
        ws = writer.sheets["Rank Stability"]

        # ── Auto-fit column widths ──
        from openpyxl.utils import get_column_letter
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3

        # ── Add metadata footer ──
        footer_row = len(df_out) + 3
        ws.cell(row=footer_row, column=1,
                value=f"Simulations: {n_sims:,}  |  Sampling: Dirichlet(1,1,1)")

    print(f"\nRank stability table exported to: {output_path}")

export_rank_stability_xlsx(city_names, all_ranks_df, OUTPUT_STABILITY_XLSX)

# ── END STEP 11 ─────────────────────────────────────────────────────────────


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  STEP 12 — OPTIONAL EXPORT: Rank Distribution Image for Focus Cities   ║
# ║  To disable: comment out everything from this banner to END STEP 12    ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def plot_rank_distributions(city_names, all_ranks_df, focus_cities, output_path,
                            X=None, base_weights=None):
    """
    For every city in focus_cities (None → all cities alphabetically), draw a
    bar chart of the rank probability distribution (x = rank position,
    y = % of simulations).  All charts are tiled into one shareable image on
    a light background.

    If X and base_weights are supplied, a vertical dashed line marks each
    city's Classic GIPI rank (the rank produced by the base weight scheme).
    """
    # ── Resolve focus list: None → all cities sorted alphabetically ───────
    if focus_cities is None:
        candidate_list = sorted(all_ranks_df.columns.tolist())
    else:
        candidate_list = focus_cities

    # ── Resolve city names (case-insensitive fallback) ────────────────────
    matched_cities = []
    for city in candidate_list:
        m = city if city in all_ranks_df.columns else next(
            (c for c in all_ranks_df.columns if c.lower() == city.lower()), None
        )
        if m is None:
            print(f"[STEP 12 WARNING] '{city}' not found in dataset — skipping.")
        else:
            matched_cities.append(m)

    if not matched_cities:
        print("[STEP 12] No valid cities found — skipping image export.")
        return

    # ── Compute Classic GIPI ranks (base weights) if provided ─────────────
    classic_ranks = {}   # city_name → int rank
    if X is not None and base_weights is not None:
        base_rank_vec = rank_cities(X, base_weights)
        for i, city in enumerate(city_names):
            classic_ranks[city] = int(base_rank_vec[i])

    n_cities  = len(city_names)
    n_focus   = len(matched_cities)
    n_sims    = len(all_ranks_df)
    rank_pos  = np.arange(1, n_cities + 1)

    # ── Shared y-axis ceiling (compute before plotting) ───────────────────
    global_max_prob = 0.0
    for city in matched_cities:
        col_data = all_ranks_df[city]
        counts   = col_data.value_counts()
        probs    = np.array([counts.get(r, 0) / n_sims * 100 for r in rank_pos])
        global_max_prob = max(global_max_prob, probs.max())
    y_max = global_max_prob * 1.15   # 15 % headroom (extra room for Classic label)

    # ── Grid layout (light thesis theme) ──────────────────────────────────
    ncols = min(4, n_focus)          # up to 4 columns for larger city sets
    nrows = int(np.ceil(n_focus / ncols))

    BG          = "white"
    BAR_CLR     = "#5c7cba"          # muted steel-blue
    ACC_CLR     = "#c0392b"          # academic red for the modal rank
    CLASSIC_CLR = "#2e7d32"          # forest green for Classic GIPI line
    TXT_CLR     = "#1a1a1a"

    fig, axes = plt.subplots(
        nrows, ncols,
        figsize=(ncols * 4.8, nrows * 3.6),
        facecolor=BG,
        squeeze=False,
    )
    fig.suptitle(
        "GIPI Rank Probability Distributions — All Cities\n"
        f"(n = {n_sims:,} Dirichlet weight samples  |  green line = Classic GIPI rank)",
        color=TXT_CLR, fontsize=14, fontweight="bold", y=1.01,
    )

    for idx, city in enumerate(matched_cities):
        row, col = divmod(idx, ncols)
        ax = axes[row][col]
        ax.set_facecolor("#f7f7f7")

        col_data  = all_ranks_df[city]
        counts    = col_data.value_counts()
        probs     = np.array([counts.get(r, 0) / n_sims * 100 for r in rank_pos])
        modal     = rank_pos[np.argmax(probs)]
        classic_r = classic_ranks.get(city, None)

        # ── Bar colours: accent the modal rank ────────────────────────────
        colors = [ACC_CLR if r == modal else BAR_CLR for r in rank_pos]

        ax.bar(rank_pos, probs, color=colors, width=0.75, zorder=3)

        # ── Classic GIPI vertical line ─────────────────────────────────────
        if classic_r is not None:
            ax.axvline(
                x=classic_r, color=CLASSIC_CLR, linewidth=1.8,
                linestyle="--", zorder=5,
            )

        ax.set_xlim(0.5, n_cities + 0.5)
        ax.set_ylim(0, y_max)          # shared scale across all panels
        ax.set_xticks(rank_pos)
        ax.set_xlabel("Rank", color=TXT_CLR, fontsize=9)
        ax.set_ylabel("Probability (%)", color=TXT_CLR, fontsize=9)
        ax.tick_params(colors=TXT_CLR, labelsize=8)
        for spine in ax.spines.values():
            spine.set_edgecolor("#ccc")
        ax.yaxis.grid(True, color="#ddd", linewidth=0.6, zorder=0)
        ax.set_axisbelow(True)

        # ── Stats annotation ──────────────────────────────────────────────
        mean_r = col_data.mean()
        std_r  = col_data.std()
        classic_note = f"  classic={classic_r}" if classic_r is not None else ""
        ax.set_title(
            f"{city}\nmean={mean_r:.1f}  σ={std_r:.2f}  mode={modal}{classic_note}",
            color=TXT_CLR, fontsize=10, fontweight="bold", pad=6,
        )

    # ── Hide unused subplots ─────────────────────────────────────────────
    for idx in range(n_focus, nrows * ncols):
        row, col = divmod(idx, ncols)
        axes[row][col].set_visible(False)

    plt.tight_layout()
    plt.savefig(output_path, dpi=180, bbox_inches="tight", facecolor=fig.get_facecolor())
    print(f"Rank distribution image saved to: {output_path}")
    plt.show()

plot_rank_distributions(city_names, all_ranks_df, FOCUS_CITIES, OUTPUT_RANKDIST_IMAGE,
                        X=X, base_weights=BASE_WEIGHTS)

# ── END STEP 12 ─────────────────────────────────────────────────────────────
