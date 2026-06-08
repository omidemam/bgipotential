# SENSITIVITY ANALYSIS — 2-Simplex Grid Search with R² vs. Base Weights (All Cities Distribution)
# This script is a sensitivity analysis of a composite index called the Green Infrastructure Potential Index,
# quantifying the potential of green infrastructure to mitigate urban heat island effect, flooding, and air pollution.

import numpy as np
import pandas as pd
import os
import matplotlib.pyplot as plt
import matplotlib.tri as mtri
import matplotlib.patheffects as pe
from matplotlib.patches import PathPatch
from matplotlib.path import Path

# ─────────────────────────────────────────────
# USER SETTINGS — edit these before running
# ─────────────────────────────────────────────
SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT    = os.path.abspath(os.path.join(SCRIPT_DIR, "..", ".."))
FILEPATH     = os.path.join(REPO_ROOT, "Results", "GIPI_Results.xls")
BASE_WEIGHTS = np.array([0.60, 0.30, 0.10])   # HydroWeight, HeatWeight, AQWeight
N_SIMS       = 5000                             # more points = smoother surface
RANDOM_SEED  = 35

# ── Optional export paths ──────────────
OUTPUT_DIR            = os.path.join(REPO_ROOT, "Visualization", "Figures and maps")
OUTPUT_STABILITY_XLSX = os.path.join(OUTPUT_DIR, "GIPI_RankStability_AllCities_6_6.xlsx")
OUTPUT_RANKDIST_IMAGE = os.path.join(OUTPUT_DIR, "GIPI_RankDistributions_AllCities_6_6.png")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Column definitions
HYDRO_COLS = ["HydroRiskAvg"]
HEAT_COLS  = ["HeatRiskAvg"]
AQ_COLS    = ["AQAvg"]


# ─────────────────────────────────────────────
# STEP 1 — Load and prepare data
# ─────────────────────────────────────────────
def load_and_prepare_data(filepath, hydro_cols, heat_cols, aq_cols):
    ext = filepath.rsplit(".", 1)[-1].lower()
    if ext in ("xls", "xlsx"):
        data = pd.read_excel(filepath)
    else:
        data = pd.read_csv(filepath)
    # Normalise compound city names (e.g. "LosAngeles" → "Los Angeles")
    CITY_NAME_MAP = {
        "LosAngeles":    "Los Angeles",
        "NewYork":       "New York",
        "NYC":           "New York City",
        "SanFrancisco":  "San Francisco",
        "SanDiego":      "San Diego",
        "SanAntonio":    "San Antonio",
        "SanJose":       "San Jose",
        "FortWorth":     "Fort Worth",
        "ElPaso":        "El Paso",
        "LasVegas":      "Las Vegas",
        "KansasCity":    "Kansas City",
        "NewOrleans":    "New Orleans",
        "SaltLakeCity":  "Salt Lake City",
        "MinneapolisSt.Paul": "Minneapolis–St. Paul",
    }
    city_names = np.array([CITY_NAME_MAP.get(c, c) for c in data["City"].values])

    hydro_score = data[hydro_cols].mean(axis=1).values
    heat_score  = data[heat_cols].mean(axis=1).values
    aq_score    = data[aq_cols].mean(axis=1).values

    X = np.column_stack([hydro_score, heat_score, aq_score])  # shape: (n_cities, 3)
    return city_names, X


# ─────────────────────────────────────────────
# STEP 2 — Compute rank vector for a given weight set
# ─────────────────────────────────────────────
def rank_cities(X, weights):
    """Score and rank cities (1 = highest score)."""
    scores = X @ weights
    return np.argsort(np.argsort(-scores)) + 1


# ─────────────────────────────────────────────
# STEP 3 — Sample weight schemes uniformly over the 2-simplex
# ─────────────────────────────────────────────
def sample_simplex(n_sims, rng):
    return rng.dirichlet([1, 1, 1], size=n_sims)


# ─────────────────────────────────────────────
# STEP 4 — Run grid search and compute R² vs. base ranking
# ─────────────────────────────────────────────
def run_grid_search(X, base_weights, weight_schemes, city_names):
    base_ranks = rank_cities(X, base_weights)
    n_cities   = X.shape[0]

    records   = []
    rank_rows = []

    for w in weight_schemes:
        w1, w2, w3 = w
        trial_ranks = rank_cities(X, w)

        r = np.corrcoef(base_ranks, trial_ranks)[0, 1]
        r_squared = r ** 2

        mean_rank_shift = np.mean(np.abs(trial_ranks - base_ranks))

        records.append({
            "w_hydro"       : round(w1, 4),
            "w_heat"        : round(w2, 4),
            "w_aq"          : round(w3, 4),
            "R2"            : round(r_squared, 6),
            "MeanRankShift" : round(mean_rank_shift, 4),
        })
        rank_rows.append(trial_ranks)

    results      = pd.DataFrame(records)
    all_ranks_df = pd.DataFrame(rank_rows, columns=city_names)

    return results.sort_values("R2", ascending=False).reset_index(drop=True), all_ranks_df


# ─────────────────────────────────────────────
# STEP 6 — Summary statistics: rank std-dev + distributions
# ─────────────────────────────────────────────
FOCUS_CITIES = None   # None → generate distributions for ALL cities in the dataset

def summary_statistics(city_names, all_ranks_df, focus_cities, X=None, base_weights=None):
    n_cities = len(city_names)
    n_sims   = len(all_ranks_df)

    if focus_cities is None:
        resolved_focus = sorted(all_ranks_df.columns.tolist())
    else:
        resolved_focus = focus_cities

    classic_ranks = {}
    if X is not None and base_weights is not None:
        base_rank_vec = rank_cities(X, base_weights)
        for i, city in enumerate(city_names):
            classic_ranks[city] = int(base_rank_vec[i])

    print("\n" + "═" * 56)
    print("  RANK STABILITY — Std-Dev Across All Weight Schemes")
    print("═" * 56)
    print(f"  (based on {n_sims:,} sampled weight triplets)\n")
    print(f"  {'City':<25} {'Mean Rank':>10} {'Median Rank':>12} {'Rank Std-Dev':>13} {'Classic GIPI':>13}")
    print("  " + "-" * 76)

    rank_stats = []
    for city in city_names:
        col = all_ranks_df[city]
        rank_stats.append({
            "city"    : city,
            "mean"    : col.mean(),
            "median"  : col.median(),
            "std"     : col.std(),
            "classic" : classic_ranks.get(city, ""),
        })
    rank_stats_df = pd.DataFrame(rank_stats).sort_values("mean")

    for _, row in rank_stats_df.iterrows():
        classic_str = f"{int(row['classic']):>13}" if row['classic'] != "" else f"{'N/A':>13}"
        print(f"  {row['city']:<25} {row['mean']:>10.2f} {row['median']:>12.1f} {row['std']:>13.4f}{classic_str}")

    bar_width = 30

    for city in resolved_focus:
        matched = city if city in all_ranks_df.columns else next(
            (c for c in all_ranks_df.columns if c.lower() == city.lower() or c.lower().startswith(city.lower())), None
        )
        if matched is None:
            print(f"\n  [WARNING] '{city}' not found in dataset — skipping distribution.")
            continue

        col       = all_ranks_df[matched]
        mean_rank = col.mean()
        med_rank  = col.median()
        std_rank  = col.std()
        classic_r = classic_ranks.get(matched, None)

        print("\n" + "═" * 56)
        print(f"  RANK PROBABILITY DISTRIBUTION — {matched}")
        print("═" * 56)
        classic_label = f"   Classic GIPI rank: {classic_r}" if classic_r is not None else ""
        print(f"  Mean rank: {mean_rank:.2f}   Median: {med_rank:.1f}   Std-Dev: {std_rank:.4f}{classic_label}")
        print(f"  {'Rank':<6} {'Count':>6} {'Prob %':>8}   Bar")
        print("  " + "-" * 56)

        counts = col.value_counts().sort_index()
        for rank_pos in range(1, n_cities + 1):
            count = counts.get(rank_pos, 0)
            prob  = count / n_sims * 100
            bar   = "█" * int(round(prob / 100 * bar_width))
            classic_marker = "  ◄ Classic GIPI" if (classic_r is not None and rank_pos == classic_r) else ""
            print(f"  {rank_pos:<6} {count:>6,} {prob:>7.2f}%   {bar}{classic_marker}")

# ─────────────────────────────────────────────
# STEP 7 — Run pipeline
# ─────────────────────────────────────────────
rng = np.random.default_rng(RANDOM_SEED)

city_names, X  = load_and_prepare_data(FILEPATH, HYDRO_COLS, HEAT_COLS, AQ_COLS)
weight_schemes = sample_simplex(N_SIMS, rng)
results, all_ranks_df = run_grid_search(X, BASE_WEIGHTS, weight_schemes, city_names)

# ─────────────────────────────────────────────
# STEP 8 — R² Summary
# ─────────────────────────────────────────────
print(f"\nBase weights — Hydro: {BASE_WEIGHTS[0]}, Heat: {BASE_WEIGHTS[1]}, AQ: {BASE_WEIGHTS[2]}  |  {N_SIMS:,} Dirichlet samples\n")
print(f"--- R² Summary ---")
print(f"Highest R²:  {results['R2'].max():.6f}  "
      f"(w_hydro={results.iloc[0]['w_hydro']}, "
      f"w_heat={results.iloc[0]['w_heat']}, "
      f"w_aq={results.iloc[0]['w_aq']})")
print(f"Lowest  R²:  {results['R2'].min():.6f}  "
      f"(w_hydro={results.iloc[-1]['w_hydro']}, "
      f"w_heat={results.iloc[-1]['w_heat']}, "
      f"w_aq={results.iloc[-1]['w_aq']})")
print(f"Mean    R²:  {results['R2'].mean():.6f}")
print(f"Median  R²:  {results['R2'].median():.6f}")

# ─────────────────────────────────────────────
# STEP 9 — Extended rank statistics
# ─────────────────────────────────────────────
summary_statistics(city_names, all_ranks_df, FOCUS_CITIES, X=X, base_weights=BASE_WEIGHTS)


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  STEP 11 — OPTIONAL EXPORT: Rank Stability Table → Excel               ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def export_rank_stability_xlsx(city_names, all_ranks_df, output_path):
    rank_stats = []
    n_sims = len(all_ranks_df)
    for city in city_names:
        col = all_ranks_df[city]
        rank_stats.append({
            "City"        : city,
            "Mean Rank"   : round(col.mean(),   4),
            "Median Rank" : round(col.median(), 4),
            "Rank Std-Dev": round(col.std(),    4),
        })
    df_out = pd.DataFrame(rank_stats).sort_values("Mean Rank").reset_index(drop=True)
    df_out.index += 1   

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_out.to_excel(writer, sheet_name="Rank Stability", index=True, index_label="#")
        ws = writer.sheets["Rank Stability"]

        from openpyxl.utils import get_column_letter
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3

        footer_row = len(df_out) + 3
        ws.cell(row=footer_row, column=1,
                value=f"Simulations: {n_sims:,}  |  Sampling: Dirichlet(1,1,1)")

    print(f"\nRank stability table exported to: {output_path}")

export_rank_stability_xlsx(city_names, all_ranks_df, OUTPUT_STABILITY_XLSX)


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  STEP 12 — Rank Distribution Image for Focus Cities                    ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def plot_rank_distributions(city_names, all_ranks_df, focus_cities, output_path,
                            X=None, base_weights=None):

    if focus_cities is None:
        candidate_list = sorted(all_ranks_df.columns.tolist())
    else:
        candidate_list = focus_cities

    matched_cities = []
    for city in candidate_list:
        m = city if city in all_ranks_df.columns else next(
            (c for c in all_ranks_df.columns if c.lower() == city.lower() or c.lower().startswith(city.lower())), None
        )
        if m is None:
            print(f"[STEP 12 WARNING] '{city}' not found in dataset — skipping.")
        else:
            matched_cities.append(m)

    if not matched_cities:
        print("[STEP 12] No valid cities found — skipping image export.")
        return

    classic_ranks = {}
    if X is not None and base_weights is not None:
        base_rank_vec = rank_cities(X, base_weights)
        for i, city in enumerate(city_names):
            classic_ranks[city] = int(base_rank_vec[i])

    n_cities  = len(city_names)
    n_focus   = len(matched_cities)
    n_sims    = len(all_ranks_df)
    rank_pos  = np.arange(1, n_cities + 1)

    global_max_prob = 0.0
    for city in matched_cities:
        col_data = all_ranks_df[city]
        counts   = col_data.value_counts()
        probs    = np.array([counts.get(r, 0) / n_sims * 100 for r in rank_pos])
        global_max_prob = max(global_max_prob, probs.max())
    y_max = global_max_prob * 1.15   

    ncols = min(4, n_focus)
    nrows = int(np.ceil(n_focus / ncols))

    BG          = "white"
    BAR_CLR     = "#5c7cba"          
    ACC_CLR     = "#c0392b"          
    CLASSIC_CLR = "#2e7d32"          
    TXT_CLR     = "#1a1a1a"

    fig, axes = plt.subplots(
        nrows, ncols,
        figsize=(ncols * 7.5, nrows * 6.0),
        facecolor=BG,
        squeeze=False,
    )

    for idx, city in enumerate(matched_cities):
        row, col = divmod(idx, ncols)
        ax = axes[row][col]
        ax.set_facecolor("#f7f7f7")

        col_data  = all_ranks_df[city]
        counts    = col_data.value_counts()
        probs     = np.array([counts.get(r, 0) / n_sims * 100 for r in rank_pos])
        modal     = rank_pos[np.argmax(probs)]
        classic_r = classic_ranks.get(city, None)

        colors = [ACC_CLR if r == modal else BAR_CLR for r in rank_pos]

        ax.bar(rank_pos, probs, color=colors, width=0.75, zorder=3)

        if classic_r is not None:
            ax.axvline(
                x=classic_r, color=CLASSIC_CLR, linewidth=2.5,
                linestyle="--", zorder=5,
            )

        ax.set_xlim(0.5, n_cities + 0.5)
        ax.set_ylim(0, y_max)          
        ax.set_xticks(rank_pos)
        ax.set_xlabel("Rank", color=TXT_CLR, fontsize=15)
        ax.set_ylabel("Probability (%)", color=TXT_CLR, fontsize=15)
        ax.tick_params(colors=TXT_CLR, labelsize=13)
        for spine in ax.spines.values():
            spine.set_edgecolor("#ccc")
        ax.yaxis.grid(True, color="#ddd", linewidth=0.8, zorder=0)
        ax.set_axisbelow(True)

        mean_r = col_data.mean()
        std_r  = col_data.std()
        if classic_r is not None:
            classic_note = f"  |  Green line = 'Classic' Rank ({classic_r})"
        else:
            classic_note = ""
        # Stats line sits just above the axes via set_title
        ax.set_title(
            f"mean={mean_r:.1f}  \u03c3={std_r:.2f}  mode={modal}{classic_note}",
            color=TXT_CLR, fontsize=11, pad=6,
        )
        # City name floats clearly above the stats line
        ax.text(
            0.5, 1.13,
            city,
            transform=ax.transAxes, ha="center", va="bottom",
            color=TXT_CLR, fontsize=22, fontweight="bold",
        )

    for idx in range(n_focus, nrows * ncols):
        row, col = divmod(idx, ncols)
        axes[row][col].set_visible(False)

    plt.tight_layout(pad=2.0)
    plt.savefig(output_path, dpi=180, bbox_inches="tight", facecolor=fig.get_facecolor())
    print(f"Rank distribution image saved to: {output_path}")
    plt.show(block=False)
    plt.pause(1)
    plt.close()

plot_rank_distributions(city_names, all_ranks_df, FOCUS_CITIES, OUTPUT_RANKDIST_IMAGE,
                        X=X, base_weights=BASE_WEIGHTS)
